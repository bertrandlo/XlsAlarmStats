import sys
from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font

from pdstats.data_import import DataImporter, DataSeries
from argparse import ArgumentParser
from pdcomponent.data_entity import DataEntity
from pdcomponent.customer import Customer
from pdcomponent.station import Station
from pdcomponent.device import Device


def main(filename):
    df = DataImporter.xls_import(filename)
    for idx, col in enumerate(df.columns):
        if col[0:7] != 'Unnamed':
            ds = DataSeries(df, idx, col)
            ds.analyze(2)
            ds.report()


def load_group_data(group_id, dt_begin, dt_end, device_instance=None, pre_evaluating_mv=5):
    from pdcomponent.device import Device
    if device_instance is None:
        device_instance = Device(group_id)

    dt_begin = datetime.strptime(dt_begin, '%Y-%m-%dT%H:%M:%S')
    dt_end = datetime.strptime(dt_end, '%Y-%m-%dT%H:%M:%S')
    if ((dt_begin-dt_end).days > 180) or ((dt_begin-dt_end).days < -180):
        raise RuntimeError("Invalid Date Range!")
    print(group_id, dt_begin, dt_end)
    device_instance.begin_datetime = dt_begin
    device_instance.ending_datetime = dt_end
    device_instance.load_trend_data()

    ds_dict = dict()

    for idx, ch in enumerate(device_instance.gChannel):
        data = list()
        pdm: DataEntity
        data.append(["{}_CH{}".format(device_instance.gName, ch), "", ""])
        data.append(["timestamp", "Magnitude", "Count"])
        for row_count, pdm in enumerate(device_instance.trend_data.get(ch)):
            data.append([pdm.bdTime, pdm.bdMV, pdm.bdCount])
        df = pd.DataFrame(data)
        ds = DataSeries(df, 0, device_name=device_instance.gName, station_name=device_instance.sName)
        ds.channel = ch
        ds.gno, ds.sno, ds.sname = device_instance.gNo, device_instance.sNo_link, device_instance.sName
        try:
            ds.analyze_by_specific_voltage(pre_evaluating_mv)
        except IndexError:
            print("Ignore.")
        ds.report()
        ds_dict[ch] = ds

    return ds_dict


def evaluating_thresholding(report_name: str, target_list: list, dt_begin, dt_end, evaluating_duration_minutes, evaluating_mv, ratio_list):
    fromFmt = "%Y-%m-%dT%H:%M:%S"
    toFmt = "%Y-%m-%dT:%H:%M:%S"
    row_offset = 1
    wb = openpyxl.Workbook()
    try:
        ws = wb["Report"]
    except KeyError:
        ws = wb.create_sheet("Report")
    # target_list = cu.get_all_devices()
    # target_list = [Device(12276)]
    dev: Device
    for dev_idx, dev in enumerate(target_list):
        print("{} / {} = {}%".format(dev_idx + 1, len(target_list),
                                     "{:.1f}".format((dev_idx + 1) / len(target_list) * 100)))
        target = dev.gNo
        ds_dict: dict
        ds_dict = load_group_data(group_id=target, dt_begin=dt_begin, dt_end=dt_end, device_instance=dev)

        ts_begin = datetime.strptime(dt_begin, fromFmt)
        ts_end = datetime.strptime(dt_end, fromFmt)

        section_title = "{}_{}_{}".format(target, ts_begin.strftime(toFmt), ts_end.strftime(toFmt))
        ws.cell(row_offset, 1).value = section_title
        ws.cell(row_offset, 8).value = dev.sName
        ws.cell(row_offset, 9).value = dev.gName
        ws.merge_cells(start_row=row_offset + 1, end_row=row_offset + 1, start_column=5,
                       end_column=(4 + 3 * len(ratio_list)))

        ws.cell(row_offset + 1, 5).value = "X STD"
        ws.cell(row_offset + 1, 5).fill = PatternFill(start_color="79DCEB", end_color="79DCEB", fill_type="solid")
        ws.cell(row_offset + 1, 5).alignment = Alignment(horizontal='center')

        ws.cell(row_offset + 2, 2).value = "CH"
        ws.cell(row_offset + 2, 2).alignment = Alignment(horizontal='center')
        ws.cell(row_offset + 2, 3).value = "Mean"
        ws.cell(row_offset + 2, 3).alignment = Alignment(horizontal='center')
        ws.cell(row_offset + 2, 4).value = "STD"
        ws.cell(row_offset + 2, 4).alignment = Alignment(horizontal='center')

        for raio_idx, ratio in enumerate(ratio_list):
            ws.cell(row_offset + 2, 5 + raio_idx * 3).value = ratio
            ws.cell(row_offset + 2, 5 + raio_idx * 3).alignment = Alignment(horizontal='center')
            ws.cell(row_offset + 2, 5 + raio_idx * 3 + 1).value = "MINs"
            ws.cell(row_offset + 2, 5 + raio_idx * 3 + 1).alignment = Alignment(horizontal='center')
            ws.cell(row_offset + 2, 5 + raio_idx * 3 + 2).value = "#"
            ws.cell(row_offset + 2, 5 + raio_idx * 3 + 2).alignment = Alignment(horizontal='center')

        for idx, channel in enumerate(dev.gChannel):
            ds = ds_dict[channel]
            ds.ratio_list = ratio_list
            ds.voltage_threshold_min = 1
            ds.lasting_minutes_min = 3
            print(ds)
            ds.report()

            ws.cell(row_offset + 3 + idx, 2).value = channel
            ws.cell(row_offset + 3 + idx, 3).value = float(f"{ds.voltage.mean():.1f}")
            ws.cell(row_offset + 3 + idx, 4).value = float(f"{ds.voltage.std():.1f}")

            stat_info = [ds.device_name + "_CH" + str(channel), f"{ds.voltage.mean():.1f}",
                         f"{ds.voltage.std():.1f}"]

            for raio_idx, ratio in enumerate(ratio_list):
                try:
                    stat_info.append(ds.report_list[str(ratio)][2])
                    stat_info.append(ds.report_list[str(ratio)][3])
                    stat_info.append(ds.report_list[str(ratio)][4])
                    ws.cell(row_offset + 3 + idx, 5 + raio_idx * 3).value = float(ds.report_list[str(ratio)][2])
                    ws.cell(row_offset + 3 + idx, 5 + raio_idx * 3).font = Font(color="FF0000")
                    ws.cell(row_offset + 3 + idx, 5 + raio_idx * 3).fill = PatternFill(start_color="EBE8F0",
                                                                                       end_color="EBE8F0",
                                                                                       fill_type="solid")
                    ws.cell(row_offset + 3 + idx, 5 + raio_idx * 3 + 1).value = int(
                        float(ds.report_list[str(ratio)][3]))
                    ws.cell(row_offset + 3 + idx, 5 + raio_idx * 3 + 2).value = float(ds.report_list[str(ratio)][4])
                except KeyError:
                    ws.cell(row_offset + 3 + idx, 5 + raio_idx * 3).value = "NaN"

        row_offset = row_offset + 6
        ws.cell(row_offset, 3).value = "觸發次數評估"
        row_offset = row_offset + 1

        ws.cell(row_offset, 3).value = "{}mv".format(evaluating_mv)

        for idx, _min_ in enumerate(evaluating_duration_minutes):
            ws.cell(row_offset, 4 + idx).value = "{}min".format(_min_)

        for idx, channel in enumerate(dev.gChannel):
            ds = ds_dict[channel]
            row_offset = row_offset + 1
            ws.cell(row_offset, 2).value = channel

            for idx, _min_ in enumerate(evaluating_duration_minutes):
                try:
                    ds.analyze_by_specific_voltage(evaluating_mv)
                    o = ds.count_occurrence(_min_)[0]
                    ws.cell(row_offset, 4 + idx).value = o
                except IndexError as e:
                    print(e)
                    ws.cell(row_offset, 4 + idx).value = 0

        row_offset = row_offset + 5
    wb.save("{}_{}.xlsx".format(report_name, datetime.now().strftime("%Y%m%dT%H%M%S")))


if __name__ == "__main__":
    parser = ArgumentParser(add_help=True)
    group_file = parser.add_argument_group()
    group_file.add_argument("--file", help="loading xlsx file exported by PDSimply")

    group_exclude = parser.add_mutually_exclusive_group(required=True)
    group_exclude.add_argument("--gno", type=int, help="loading data from PUMO database by the specified group id")
    group_exclude.add_argument("--sno", type=int, help="loading data from PUMO database by the specified station id")
    group_exclude.add_argument("--cuno", type=int, help="loading data from PUMO database by the specified customer id")

    group_loading = parser.add_argument_group()
    group_loading.add_argument("--begin", type=str, help="begin datetime, max range 180 days., %Y-%m-%dT%H:%M:%S")
    group_loading.add_argument("--end", type=str, help="end datetime, %Y-%m-%dT%H:%M:%S")

    group_minutes_list = parser.add_argument_group()
    group_minutes_list.add_argument("--minutes", nargs="+", type=int,
                               help="minutes set for evaluating trigger count, default 10 15 20 25 30")

    group_ratio_list = parser.add_argument_group()
    group_ratio_list.add_argument("--ratio", nargs="+", type=float,
                               help="std ratio set for evaluating thresholding count, default 1.0 2.0 3.0")

    group_evaluating_mv = parser.add_argument_group()
    group_evaluating_mv.add_argument("--eval_mv", type=int, help="mv for evaluating trigger count, default 30")

    args = parser.parse_args()

    dt_begin = args.begin
    dt_end = args.end
    ratio_list = [1, 2, 3]  # --ratio
    evaluating_duration_minutes = [10, 15, 20, 25, 30]  # --minutes
    eval_mv = 30  # --eval_mv

    if args.ratio is not None:
        ratio_list = list(args.ratio)
        ratio_list.sort()

    print("Ratio List = {}".format(ratio_list))

    if args.minutes is not None:
        evaluating_duration_minutes = list(args.minutes)
        evaluating_duration_minutes.sort()

    print("Eval Minutes List = {}".format(evaluating_duration_minutes))

    if args.eval_mv is not None:
        eval_mv = int(args.eval_mv)

    print("Eval mV = {}".format(eval_mv))

    if args.gno is not None:
        evaluating_thresholding("Evaluating_Report", [Device(args.gno)],
                                dt_begin=dt_begin, dt_end=dt_end,
                                evaluating_duration_minutes=evaluating_duration_minutes, evaluating_mv=eval_mv,
                                ratio_list=ratio_list)
        sys.exit()

    if args.sno is not None:
        st = Station(sNo=args.sno)
        evaluating_thresholding("Evaluating_Report", st.devices,
                                dt_begin=dt_begin, dt_end=dt_end,
                                evaluating_duration_minutes=evaluating_duration_minutes, evaluating_mv=eval_mv,
                                ratio_list=ratio_list)
        sys.exit()

    if args.cuno is not None:
        cu = Customer(cuNo=args.cuno)
        evaluating_thresholding("Evaluating_Report", cu.get_all_devices(),
                                dt_begin=dt_begin, dt_end=dt_end,
                                evaluating_duration_minutes=evaluating_duration_minutes, evaluating_mv=eval_mv,
                                ratio_list=ratio_list)
        sys.exit()

    if args.file is not None:
        main(filename=args.file)
        sys.exit()

    parser.print_help()
